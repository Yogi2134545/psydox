"""Nano Banana — main orchestrator engine."""
import io
import traceback
from pathlib import Path
from typing import Callable, Optional

from PIL import Image

from .background_generator import BackgroundGenerator
from .lifestyle_generator import LifestyleGenerator
from .model_generator import ModelGenerator
from .ai_editor import AIEditor
from .product_enhancer import ProductEnhancer
from .prompt_builder import (
    build_scene_prompt,
    build_lighting_prompt,
    build_shadow_prompt,
)
from .api_client import GeminiClient
from .export import Exporter


class NanoBananaEngine:
    def __init__(self):
        self.bg_gen = BackgroundGenerator()
        self.lifestyle_gen = LifestyleGenerator()
        self.model_gen = ModelGenerator()
        self.editor = AIEditor()
        self.enhancer = ProductEnhancer()
        self.client = GeminiClient()
        self.exporter = Exporter()

    def process_single(self, image: Image.Image, config: dict) -> Image.Image:
        """
        Route a single image to the appropriate generator.

        config keys:
          mode: "background" | "lifestyle" | "model" | "edit" | "enhance" | "scene"
                "lighting" | "shadow"
          ... mode-specific params ...
        """
        mode = config.get("mode", "background")

        if mode == "background":
            return self.bg_gen.replace_background(
                image,
                config.get("background_option", "White"),
                config.get("custom_prompt", ""),
                config.get("product_desc", ""),
            )

        elif mode == "lifestyle":
            return self.lifestyle_gen.generate(
                image,
                config.get("style", "Casual Street Style"),
                config.get("custom_prompt", ""),
                config.get("product_desc", ""),
            )

        elif mode == "model":
            return self.model_gen.generate(
                image,
                config.get("gender", "Female"),
                config.get("age_group", "25-35"),
                config.get("ethnicity", "South Asian / Indian"),
                config.get("clothing_style", "Natural / Minimal"),
                config.get("product_desc", ""),
            )

        elif mode == "edit":
            settings = config.get("settings", {})
            result = self.editor.adjust(image, settings)
            ai_finish = config.get("ai_finish", "None")
            if ai_finish and ai_finish != "None":
                result = self.editor.apply_ai_finish(result, ai_finish)
            return result

        elif mode == "enhance":
            return self.enhancer.enhance(
                image,
                config.get("enhancements", []),
                config.get("product_desc", ""),
            )

        elif mode == "scene":
            return self._generate_scene(image, config)

        elif mode == "lighting":
            return self._apply_lighting(image, config)

        elif mode == "shadow":
            return self._apply_shadow(image, config)

        else:
            return image

    _NO_API_MSG = (
        "This feature requires Imagen API access.\n\n"
        "Your GOOGLE_API_KEY supports text + vision only.\n"
        "To enable: go to https://aistudio.google.com and enable Imagen 3 (requires billing).\n\n"
        "Background tab works without Imagen."
    )

    def _generate_scene(self, image: Image.Image, config: dict) -> Image.Image:
        prompt = build_scene_prompt(config.get("scene_type", "Hero Product Shot"), config.get("product_desc", ""))
        ref_bytes = io.BytesIO()
        image.save(ref_bytes, format="PNG")
        try:
            result_bytes = self.client.generate_image(prompt, reference_image_bytes=ref_bytes.getvalue())
            return Image.open(io.BytesIO(result_bytes)).convert("RGB")
        except Exception as e:
            if any(c in str(e) for c in ("404", "400", "403")):
                raise RuntimeError(self._NO_API_MSG) from None
            raise

    def _apply_lighting(self, image: Image.Image, config: dict) -> Image.Image:
        prompt = build_lighting_prompt(config.get("lighting_type", "Soft Studio"), config.get("product_desc", ""))
        img_bytes = io.BytesIO()
        image.save(img_bytes, format="PNG")
        try:
            result_bytes = self.client.edit_image(img_bytes.getvalue(), prompt)
            return Image.open(io.BytesIO(result_bytes)).convert("RGB")
        except Exception as e:
            if any(c in str(e) for c in ("404", "400", "403")):
                raise RuntimeError(self._NO_API_MSG) from None
            raise

    def _apply_shadow(self, image: Image.Image, config: dict) -> Image.Image:
        prompt = build_shadow_prompt(config.get("shadow_type", "Natural Drop Shadow"))
        img_bytes = io.BytesIO()
        image.save(img_bytes, format="PNG")
        try:
            result_bytes = self.client.edit_image(img_bytes.getvalue(), prompt)
            return Image.open(io.BytesIO(result_bytes)).convert("RGB")
        except Exception as e:
            if any(c in str(e) for c in ("404", "400", "403")):
                raise RuntimeError(self._NO_API_MSG) from None
            raise

    # ── Batch processing ──────────────────────────────────────────────────────

    def process_batch(
        self,
        excel_path: str,
        config: dict,
        progress_cb: Optional[Callable] = None,
    ) -> dict:
        """
        Process images from an Excel file (same format as classic Psydox).
        Returns results dict compatible with existing ZIP export format.
        """
        import requests
        import openpyxl
        import zipfile
        import tempfile

        results = {"total": 0, "success": 0, "failed": 0, "skipped": 0}
        processed_images = []

        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

            rows = list(ws.iter_rows(min_row=2, values_only=True))
            results["total"] = len(rows)

            for i, row in enumerate(rows):
                row_dict = dict(zip(headers, row))
                style_code = str(row_dict.get("STYLE_CODE", f"item_{i+1}")).strip()

                # Find image URLs (IMAGE1..IMAGE12)
                img_urls = []
                for j in range(1, 13):
                    url = row_dict.get(f"IMAGE{j}")
                    if url and str(url).strip().startswith("http"):
                        img_urls.append(str(url).strip())

                if not img_urls:
                    results["skipped"] += 1
                    if progress_cb:
                        progress_cb(i + 1, results["total"])
                    continue

                n_angles = config.get("angles", 1)
                # Process first image (or all, configurable)
                for url in img_urls[:1]:
                    try:
                        resp = requests.get(url, timeout=30)
                        resp.raise_for_status()
                        src_img = Image.open(io.BytesIO(resp.content)).convert("RGB")
                        if n_angles <= 1:
                            result_img = self.process_single(src_img, config)
                            processed_images.append((f"{style_code}.jpg", result_img))
                            results["success"] += 1
                        else:
                            ref_buf = io.BytesIO()
                            src_img.convert("RGB").save(ref_buf, format="JPEG", quality=90)
                            bg_opt = config.get("background_option", "White")
                            base_prompt = f"Product photo with {bg_opt} background"
                            angle_bytes_list = self.client.generate_angles(
                                base_prompt, ref_buf.getvalue(), count=n_angles
                            )
                            for ai, ab in enumerate(angle_bytes_list):
                                if ab is not None:
                                    angle_img = Image.open(io.BytesIO(ab)).convert("RGB")
                                    processed_images.append(
                                        (f"{style_code}_angle{ai + 1}.jpg", angle_img)
                                    )
                                    results["success"] += 1
                                else:
                                    results["failed"] += 1
                    except Exception:
                        results["failed"] += 1

                if progress_cb:
                    progress_cb(i + 1, results["total"])

        except Exception as e:
            results["error"] = str(e)
            return results

        # Build ZIP
        if processed_images:
            zip_buf = self.exporter.batch_to_zip(processed_images, fmt="JPEG", quality=90)
            # Save to temp file
            tmp = Path(tempfile.mkdtemp()) / "nb_batch_output.zip"
            tmp.write_bytes(zip_buf)
            results["zip_path"] = str(tmp)
            results["zip_bytes"] = zip_buf

        return results
