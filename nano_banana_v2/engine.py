"""
Nano Banana v2 — improved orchestrator engine.

Improvements over v1:
  - Uses v2 prompt_builder: richer prompts, fidelity in all types, aspect hint
  - Transparent-image safe handling: RGBA images are composited onto neutral grey
    before being passed to the API (avoids API artefacts from raw alpha channels)
  - Single JPEG quality constant: change EXPORT_QUALITY once, applies everywhere
  - Fixed error messages: no longer says "Imagen 3" — says "Imagen API"
  - Uses AIEditorV2 (fixed shadows curve, no dead GeminiClient)
  - All other generators imported directly from nano_banana (no duplication)
"""
from __future__ import annotations

import io
import logging
from pathlib import Path
from typing import Callable, Optional

_log = logging.getLogger("nano_banana_v2.engine")

from PIL import Image

# Reuse unchanged generators and client from nano_banana
from nano_banana.background_generator import BackgroundGenerator
from nano_banana.lifestyle_generator  import LifestyleGenerator
from nano_banana.model_generator      import ModelGenerator
from nano_banana.product_enhancer     import ProductEnhancer
from nano_banana.api_client           import GeminiClient
from nano_banana.export               import Exporter
from nano_banana.validators           import validate_url, URLValidationError

# v2 replacements
from .ai_editor     import AIEditorV2
from .prompt_builder import (
    build_scene_prompt,
    build_lighting_prompt,
    build_shadow_prompt,
)

EXPORT_QUALITY = 90  # single constant for all JPEG exports


class NanoBananaEngineV2:
    """Improved orchestrator — same interface as NanoBananaEngine but better prompts."""

    def __init__(self):
        self.bg_gen       = BackgroundGenerator()
        self.lifestyle_gen = LifestyleGenerator()
        self.model_gen    = ModelGenerator()
        self.editor       = AIEditorV2()
        self.enhancer     = ProductEnhancer()
        self.client       = GeminiClient()
        self.exporter     = Exporter()

    # ── Single image processing ───────────────────────────────────────────────

    def process_single(self, image: Image.Image, config: dict) -> Image.Image:
        """
        Route a single image to the appropriate generator.

        config keys:
          mode          : "background" | "lifestyle" | "model" | "edit" | "enhance"
                          "scene" | "lighting" | "shadow"
          ratio_label   : (optional) string like "1:1", "4:5", "9:16" — added as
                          composition hint in prompts
          ... mode-specific params (same as v1) ...
        """
        mode        = config.get("mode", "background")
        ratio_label = config.get("ratio_label", "")

        if mode == "background":
            return self.bg_gen.replace_background(
                image,
                config.get("background_option", "White"),
                config.get("custom_prompt", ""),
                config.get("product_desc", ""),
            )

        elif mode == "lifestyle":
            return self.lifestyle_gen.generate(
                image,
                config.get("style", "Casual Street Style"),
                config.get("custom_prompt", ""),
                config.get("product_desc", ""),
            )

        elif mode == "model":
            return self.model_gen.generate(
                image,
                config.get("gender", "Female"),
                config.get("age_group", "25-35"),
                config.get("ethnicity", "South Asian / Indian"),
                config.get("clothing_style", "Natural / Minimal"),
                config.get("product_desc", ""),
            )

        elif mode == "edit":
            result = self.editor.adjust(image, config.get("settings", {}))
            ai_finish = config.get("ai_finish", "None")
            if ai_finish and ai_finish != "None":
                result = self.editor.apply_ai_finish(result, ai_finish)
            return result

        elif mode == "enhance":
            return self.enhancer.enhance(
                image,
                config.get("enhancements", []),
                config.get("product_desc", ""),
            )

        elif mode == "scene":
            return self._generate_scene(image, config, ratio_label)

        elif mode == "lighting":
            return self._apply_lighting(image, config, ratio_label)

        elif mode == "shadow":
            return self._apply_shadow(image, config, ratio_label)

        else:
            return image

    # ── Internal API-backed modes ─────────────────────────────────────────────

    _NO_API_MSG = (
        "This feature requires Imagen API access.\n\n"
        "Your GOOGLE_API_KEY supports text and vision generation only.\n"
        "To enable image editing: go to https://aistudio.google.com and enable "
        "Imagen API access (billing required).\n\n"
        "Background tab works without Imagen API access."
    )

    def _safe_png_bytes(self, image: Image.Image) -> bytes:
        """
        Return PNG bytes safe to send to the API.

        If the image has an alpha channel, composite it onto neutral grey first.
        Raw RGBA sent to the API can produce artifacts or artefact-filled results
        depending on how the API interprets semi-transparent pixels.
        """
        if image.mode == "RGBA":
            bg = Image.new("RGB", image.size, (200, 200, 200))
            bg.paste(image, mask=image.split()[3])
            img_out = bg
        else:
            img_out = image.convert("RGB")
        buf = io.BytesIO()
        img_out.save(buf, format="PNG")
        return buf.getvalue()

    def _generate_scene(
        self, image: Image.Image, config: dict, ratio_label: str
    ) -> Image.Image:
        prompt = build_scene_prompt(
            config.get("scene_type", "Hero Product Shot"),
            config.get("product_desc", ""),
            config.get("custom_prompt", ""),
            ratio_label,
        )
        try:
            result_bytes = self.client.generate_image(
                prompt, reference_image_bytes=self._safe_png_bytes(image)
            )
            return Image.open(io.BytesIO(result_bytes)).convert("RGB")
        except Exception as e:
            if any(c in str(e) for c in ("404", "400", "403")):
                raise RuntimeError(self._NO_API_MSG) from None
            raise

    def _apply_lighting(
        self, image: Image.Image, config: dict, ratio_label: str
    ) -> Image.Image:
        prompt = build_lighting_prompt(
            config.get("lighting_type", "Soft Studio"),
            config.get("product_desc", ""),
            config.get("custom_prompt", ""),
            ratio_label,
        )
        try:
            result_bytes = self.client.edit_image(self._safe_png_bytes(image), prompt)
            return Image.open(io.BytesIO(result_bytes)).convert("RGB")
        except Exception as e:
            if any(c in str(e) for c in ("404", "400", "403")):
                raise RuntimeError(self._NO_API_MSG) from None
            raise

    def _apply_shadow(
        self, image: Image.Image, config: dict, ratio_label: str
    ) -> Image.Image:
        prompt = build_shadow_prompt(
            config.get("shadow_type", "Natural Drop Shadow"),
            config.get("product_desc", ""),
            config.get("custom_prompt", ""),
            ratio_label,
        )
        try:
            result_bytes = self.client.edit_image(self._safe_png_bytes(image), prompt)
            return Image.open(io.BytesIO(result_bytes)).convert("RGB")
        except Exception as e:
            if any(c in str(e) for c in ("404", "400", "403")):
                raise RuntimeError(self._NO_API_MSG) from None
            raise

    # ── Batch processing ──────────────────────────────────────────────────────

    def process_batch(
        self,
        excel_path: str,
        config: dict,
        progress_cb: Optional[Callable] = None,
    ) -> dict:
        """
        Process images from an Excel file.
        Same interface as NanoBananaEngine.process_batch.
        """
        import requests
        import openpyxl

        results = {"total": 0, "success": 0, "failed": 0, "skipped": 0}
        processed_images = []

        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            ws = wb.active
            headers = [
                str(c.value).strip() if c.value else ""
                for c in next(ws.iter_rows(min_row=1, max_row=1))
            ]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            results["total"] = len(rows)

            for i, row in enumerate(rows):
                row_dict   = dict(zip(headers, row))
                style_code = str(row_dict.get("STYLE_CODE", f"item_{i+1}")).strip()

                img_urls = [
                    str(row_dict.get(f"IMAGE{j}", "")).strip()
                    for j in range(1, 13)
                    if row_dict.get(f"IMAGE{j}") and str(row_dict[f"IMAGE{j}"]).strip().startswith("http")
                ]

                if not img_urls:
                    results["skipped"] += 1
                    if progress_cb:
                        progress_cb(i + 1, results["total"])
                    continue

                n_angles = config.get("angles", 1)
                for url in img_urls[:1]:
                    try:
                        try:
                            url = validate_url(url)
                        except URLValidationError as ve:
                            _log.warning("Batch SSRF block: %s — url=%s", ve, url[:100])
                            results["failed"] += 1
                            if progress_cb:
                                progress_cb(i + 1, results["total"])
                            continue

                        resp    = requests.get(url, timeout=30)
                        resp.raise_for_status()
                        src_img = Image.open(io.BytesIO(resp.content)).convert("RGB")

                        if n_angles <= 1:
                            result_img = self.process_single(src_img, config)
                            processed_images.append((f"{style_code}.jpg", result_img))
                            results["success"] += 1
                        else:
                            ref_buf = io.BytesIO()
                            src_img.save(ref_buf, format="JPEG", quality=EXPORT_QUALITY)
                            bg_opt    = config.get("background_option", "White")
                            base_prompt = f"Product photo with {bg_opt} background"
                            angle_bytes_list = self.client.generate_angles(
                                base_prompt, ref_buf.getvalue(), count=n_angles
                            )
                            for ai, ab in enumerate(angle_bytes_list):
                                if ab is not None:
                                    angle_img = Image.open(io.BytesIO(ab)).convert("RGB")
                                    processed_images.append(
                                        (f"{style_code}_angle{ai+1}.jpg", angle_img)
                                    )
                                    results["success"] += 1
                                else:
                                    results["failed"] += 1
                    except Exception:
                        _log.exception("Batch row %d failed", i)
                        results["failed"] += 1

                if progress_cb:
                    progress_cb(i + 1, results["total"])

        except Exception as e:
            results["error"] = str(e)
            return results

        if processed_images:
            import tempfile
            zip_buf = self.exporter.batch_to_zip(
                processed_images, fmt="JPEG", quality=EXPORT_QUALITY
            )
            tmp = Path(tempfile.mkdtemp()) / "nb_v2_batch_output.zip"
            tmp.write_bytes(zip_buf)
            results["zip_path"]  = str(tmp)
            results["zip_bytes"] = zip_buf

        return results
