"""Psydox Auth — AuthService (main entry point for all auth operations).

AuthService is the single API surface used by UI code.
Business logic lives here; UI code only calls AuthService methods.

Thread-safety: all state is in the DB; no mutable in-process state.
"""
from __future__ import annotations

import logging
import os
import time
from typing import Optional

from psydox.auth.models import AccountStatus, AuthResult, User
from psydox.auth.password import PasswordService
from psydox.auth.tokens import TokenService
from psydox.auth.validation import normalize_email, validate_email, validate_name

_log = logging.getLogger("psydox.auth.service")

_VERIFICATION_TTL = 48 * 3600
_RESET_TTL        =  1 * 3600

# All roles that can be assigned through the admin role-management interface.
# "owner" is intentionally absent — it is a bootstrap-only role, never assignable via UI.
_VALID_ROLES = frozenset({
    "user", "viewer", "editor", "manager", "reviewer",
    "creative", "operator", "catalog_operator", "admin",
})


class AuthService:
    """
    Central auth service.  All UI code should call this, never the
    repository or sub-services directly.
    """

    def __init__(self) -> None:
        from psydox.auth.repository import get_user_repository
        from psydox.auth.sessions   import get_session_service
        self._repo     = get_user_repository()
        self._sessions = get_session_service()
        self._pw       = PasswordService()
        self._tok      = TokenService()
        self._migrated = False
        self._ensure_migrated()

    # ── Startup ───────────────────────────────────────────────────────────────

    def _ensure_migrated(self) -> None:
        """Migrate users.yaml to DB on first call (idempotent)."""
        if self._migrated:
            return
        try:
            self._repo.migrate_from_yaml()
        except Exception as exc:
            _log.warning("users.yaml migration failed (non-fatal): %s", exc)
        self._ensure_system_owner()
        self._sync_yaml_roles()
        self._migrated = True

    def _sync_yaml_roles(self) -> None:
        """Update roles for existing non-owner users to match users.yaml.

        migrate_from_yaml() only inserts NEW rows; this method handles
        the case where roles in yaml changed after users were already in the DB.
        """
        import yaml
        from pathlib import Path
        path = Path("users.yaml")
        if not path.exists():
            return
        try:
            data = yaml.safe_load(path.read_text()) or {}
        except Exception as exc:
            _log.warning("_sync_yaml_roles: could not read users.yaml: %s", exc)
            return

        for email, udata in data.items():
            if not isinstance(udata, dict):
                continue
            norm = normalize_email(email)
            target_role = udata.get("role", "operator")
            try:
                user = self._repo.get_by_email(norm)
                if user is None or user.role == "owner":
                    continue
                if user.role != target_role:
                    self._repo.update_role(user.id, target_role)
                    _log.info("_sync_yaml_roles: %s %s → %s", norm, user.role, target_role)
            except Exception as exc:
                _log.warning("_sync_yaml_roles: failed for %s: %s", norm, exc)

    def _ensure_system_owner(self) -> None:
        """Guarantee yogeshwar@popclub.co always has the 'owner' role in DB.

        On a fresh DB (e.g. Railway deploy with no persistent volume), the owner
        account does not exist yet. If PSYDOX_OWNER_PASSWORD is set in env, the
        account is auto-created so the owner can log in immediately without
        going through the registration flow.
        """
        from psydox.access import OWNER_EMAIL
        try:
            user = self._repo.get_by_email(OWNER_EMAIL)

            if user is None:
                # Fresh DB — try to seed the owner account from env var
                seed_pw = os.environ.get("PSYDOX_OWNER_PASSWORD", "").strip()
                if seed_pw:
                    pw_hash = self._pw.hash(seed_pw)
                    user = self._repo.create(
                        full_name="Yogeshwar",
                        email=OWNER_EMAIL,
                        password_hash=pw_hash,
                        role="owner",
                        email_verified=True,
                        status=AccountStatus.ACTIVE,
                    )
                    _log.info("Owner account auto-created from PSYDOX_OWNER_PASSWORD")
                else:
                    _log.warning(
                        "Owner account missing from DB and PSYDOX_OWNER_PASSWORD not set. "
                        "Register manually or set PSYDOX_OWNER_PASSWORD in Railway env vars."
                    )
                return

            # Account exists — ensure correct role and active status
            if user.role != "owner":
                self._repo.update_role(user.id, "owner")
                _log.info("Promoted %s to owner role", OWNER_EMAIL)
            if not user.email_verified or user.status == AccountStatus.PENDING_VERIFICATION:
                self._repo.verify_email(user.id)
                _log.info("Auto-verified owner email for %s", OWNER_EMAIL)

        except Exception as exc:
            _log.warning("_ensure_system_owner failed (non-fatal): %s", exc)

    # ── Registration ──────────────────────────────────────────────────────────

    def register(
        self,
        full_name:        str,
        email:            str,
        password:         str,
        confirm_password: str,
        terms_accepted:   bool = False,
    ) -> AuthResult:
        errors: list[str] = []

        name_errs = validate_name(full_name)
        errors.extend(name_errs)

        email_errs = validate_email(email)
        errors.extend(email_errs)

        pw_errs = self._pw.validate_strength(password)
        errors.extend(pw_errs)

        if password != confirm_password:
            errors.append("Passwords do not match.")

        if not terms_accepted:
            errors.append("You must accept the Terms & Privacy Policy.")

        if errors:
            return AuthResult.fail(errors[0], "VALIDATION_ERROR")

        norm_email = normalize_email(email)

        # Duplicate check — return safe generic message to avoid enumeration
        if self._repo.email_exists(norm_email):
            return AuthResult.fail(
                "An account with this email already exists. Please sign in or use a different email.",
                "DUPLICATE_EMAIL",
            )

        from psydox.access import OWNER_EMAIL
        is_owner_reg = (norm_email == normalize_email(OWNER_EMAIL))

        pw_hash = self._pw.hash(password)
        now = time.time()
        user = self._repo.create(
            full_name=full_name.strip(),
            email=norm_email,
            password_hash=pw_hash,
            role="owner" if is_owner_reg else "user",
            email_verified=is_owner_reg,
            status=AccountStatus.ACTIVE if is_owner_reg else AccountStatus.PENDING_VERIFICATION,
            terms_accepted_at=now,
        )

        _append_registration_log(user)
        self._audit(norm_email, "user_registered")
        _log.info("New user registered: %s  (owner_bootstrap=%s)", norm_email, is_owner_reg)

        if is_owner_reg:
            # Owner is immediately active — no email verification needed.
            return AuthResult.ok(user)

        raw_token = self._send_verification(user)

        # When no email service is configured, surface the link in the UI
        # so users aren't stuck waiting for a mail that never arrives.
        in_app = ""
        if raw_token and _is_console_email():
            from psydox.auth.email import _PSYDOX_BASE_URL
            in_app = f"{_PSYDOX_BASE_URL.rstrip('/')}?verify={raw_token}"

        return AuthResult.ok(user, in_app_link=in_app)

    # ── Login ─────────────────────────────────────────────────────────────────

    def login(
        self,
        email:       str,
        password:    str,
        remember_me: bool = False,
    ) -> AuthResult:
        norm_email = normalize_email(email)

        # Per-email rate limit (in-memory token bucket)
        if not self._check_rate_limit(norm_email):
            return AuthResult.fail(
                "Too many login attempts. Please wait a few minutes and try again.",
                "RATE_LIMITED",
            )

        # Generic error message — never reveal whether account exists
        _GENERIC = "Invalid email or password."

        from psydox.access import OWNER_EMAIL
        _is_owner_login = (norm_email == normalize_email(OWNER_EMAIL))

        if _is_owner_login:
            user = self._repo.get_by_email(norm_email)
            # Bootstrap if no account OR account seeded from users.yaml with empty/invalid hash
            _needs_bootstrap = user is None or not (user.password_hash or "").startswith("$2b$")
            if _needs_bootstrap:
                # First-ever login or yaml-migrated account with empty hash.
                # Set the password hash from the submitted credentials.
                try:
                    pw_hash = self._pw.hash(password)
                    if user is None:
                        self._repo.create(
                            full_name="Yogeshwar",
                            email=norm_email,
                            password_hash=pw_hash,
                            role="owner",
                            email_verified=True,
                            status=AccountStatus.ACTIVE,
                        )
                    else:
                        # Account exists but has no usable password — set it now
                        self._repo.update_password_hash(user.id, pw_hash)
                        if user.role != "owner":
                            self._repo.update_role(user.id, "owner")
                        if not user.email_verified \
                                or user.status == AccountStatus.PENDING_VERIFICATION:
                            self._repo.verify_email(user.id)
                    _log.info("Owner account bootstrapped from first login attempt")
                except Exception as _boot_err:
                    _log.error("Owner bootstrap failed: %s", _boot_err)
                    return AuthResult.fail(
                        "Account setup failed. Please use 'Create account' to register first.",
                        "BOOTSTRAP_FAILED",
                    )
            elif user.role != "owner" or not user.email_verified \
                    or user.status == AccountStatus.PENDING_VERIFICATION:
                # Account exists with a real hash but wrong state — fix it before the password check.
                self._ensure_system_owner()

        user = self._repo.get_by_email(norm_email)
        if user is None:
            self._audit(norm_email, "login_failed", detail="unknown email")
            return AuthResult.fail(_GENERIC, "INVALID_CREDENTIALS")

        # SEC-001 FIX: Non-owner users seeded from users.yaml arrive with password_hash="".
        # We must NOT accept the first submitted password as their permanent password —
        # doing so lets any attacker who knows a victim email permanently own that account.
        # Instead, block login and require an admin to set the password via a secure flow.
        if not _is_owner_login and not (user.password_hash or "").startswith("$2b$"):
            self._audit(norm_email, "login_failed", detail="account not activated (empty password_hash)")
            _log.warning("Login blocked for %s: account not activated (yaml-seeded, no password set)", norm_email)
            return AuthResult.fail(
                "Account not yet activated. Please contact the administrator to set your password.",
                "NOT_ACTIVATED",
            )

        # Account state checks
        if user.status == AccountStatus.DISABLED:
            return AuthResult.fail(
                "This account has been disabled. Please contact support.", "DISABLED"
            )
        if user.status == AccountStatus.SUSPENDED:
            return AuthResult.fail(
                "This account is currently suspended. Please contact support.", "SUSPENDED"
            )

        # Lockout check
        if user.is_locked():
            remaining = int((user.locked_until or 0) - time.time())
            mins = max(1, remaining // 60)
            return AuthResult.fail(
                f"Account is temporarily locked. Try again in {mins} minute(s).", "LOCKED"
            )

        # Password check
        if not self._pw.verify(password, user.password_hash):
            self._repo.record_failed_login(user.id)
            self._audit(norm_email, "login_failed", detail="wrong password")
            return AuthResult.fail(_GENERIC, "INVALID_CREDENTIALS")

        # Unverified — block login, let UI offer resend
        if not user.email_verified or user.status == AccountStatus.PENDING_VERIFICATION:
            return AuthResult.fail(
                "Please verify your email before signing in.",
                "UNVERIFIED",
            )

        # Success
        session = self._sessions.create(user.id, remember_me=remember_me)
        self._repo.record_login(user.id)
        self._audit(norm_email, "login_success")
        return AuthResult.ok(user, session_id=session.id)

    # ── Logout ────────────────────────────────────────────────────────────────

    def logout(self, session_id: str, user_email: str = "") -> None:
        try:
            self._sessions.revoke(session_id)
            self._audit(user_email, "logout")
        except Exception as exc:
            _log.warning("logout failed: %s", exc)

    def logout_all(self, user_id: str, user_email: str = "") -> int:
        count = self._sessions.revoke_all(user_id)
        self._audit(user_email, "logout_all_sessions", detail=f"revoked {count} sessions")
        return count

    # ── Email verification ────────────────────────────────────────────────────

    def verify_email(self, token: str) -> AuthResult:
        token = token.strip()
        if not token:
            return AuthResult.fail("Invalid verification link.", "INVALID_TOKEN")

        token_hash = self._tok.hash(token)
        db = self._db()
        row = db.execute(
            "SELECT * FROM users WHERE verification_token_hash=?", (token_hash,)
        ).fetchone()

        if not row:
            return AuthResult.fail(
                "Verification link is invalid or has already been used.", "INVALID_TOKEN"
            )

        from psydox.auth.models import User, AccountStatus
        user = self._repo._row_to_user(row)

        if user.verification_expires_at and user.verification_expires_at < time.time():
            return AuthResult.fail(
                "Verification link has expired. Please request a new one.", "TOKEN_EXPIRED"
            )

        self._repo.verify_email(user.id)
        self._audit(user.email, "email_verified")
        user = self._repo.get_by_id(user.id)
        _log.info("Email verified: %s", user.email)
        return AuthResult.ok(user)

    def resend_verification(self, email: str) -> str:
        """Resend verification email. Returns in-app link when console mode, else ''."""
        norm = normalize_email(email)
        user = self._repo.get_by_email(norm)
        in_app = ""
        if user and not user.email_verified:
            raw_token = self._send_verification(user)
            if raw_token and _is_console_email():
                from psydox.auth.email import _PSYDOX_BASE_URL
                in_app = f"{_PSYDOX_BASE_URL.rstrip('/')}?verify={raw_token}"
        self._audit(norm, "verification_resend_requested")
        return in_app

    # ── Password reset ────────────────────────────────────────────────────────

    def request_password_reset(self, email: str) -> None:
        """
        Request a password reset.
        ALWAYS returns — never reveals whether account exists.
        """
        norm = normalize_email(email)
        user = self._repo.get_by_email(norm)
        self._audit(norm, "password_reset_requested")
        if not user or user.status not in (AccountStatus.ACTIVE, AccountStatus.PENDING_VERIFICATION):
            return  # silent — do not reveal account existence
        try:
            raw_token = self._tok.generate()
            token_hash = self._tok.hash(raw_token)
            self._repo.set_reset_token(user.id, token_hash, time.time() + _RESET_TTL)
            email_svc = _get_email_service()
            email_svc.send_password_reset_email(user.email, user.full_name, raw_token)
            self._audit(norm, "password_reset_email_sent")
        except Exception as exc:
            _log.error("password reset email failed for %s: %s", norm, exc)

    def reset_password(
        self,
        token:            str,
        new_password:     str,
        confirm_password: str,
    ) -> AuthResult:
        token = token.strip()
        if not token:
            return AuthResult.fail("Invalid reset link.", "INVALID_TOKEN")

        errors = self._pw.validate_strength(new_password)
        if errors:
            return AuthResult.fail(errors[0], "WEAK_PASSWORD")
        if new_password != confirm_password:
            return AuthResult.fail("Passwords do not match.", "PASSWORD_MISMATCH")

        token_hash = self._tok.hash(token)
        db = self._db()
        row = db.execute(
            "SELECT * FROM users WHERE reset_token_hash=?", (token_hash,)
        ).fetchone()

        if not row:
            return AuthResult.fail(
                "Reset link is invalid or has already been used.", "INVALID_TOKEN"
            )

        user = self._repo._row_to_user(row)

        if user.reset_expires_at and user.reset_expires_at < time.time():
            return AuthResult.fail(
                "Reset link has expired. Please request a new one.", "TOKEN_EXPIRED"
            )

        new_hash = self._pw.hash(new_password)
        self._repo.update_password(user.id, new_hash)
        # Invalidate all sessions on password reset
        self._sessions.revoke_all(user.id)
        self._audit(user.email, "password_reset_completed")
        user = self._repo.get_by_id(user.id)
        _log.info("Password reset completed: %s", user.email)
        return AuthResult.ok(user)

    # ── Change password ───────────────────────────────────────────────────────

    def change_password(
        self,
        user_id:            str,
        current_password:   str,
        new_password:       str,
        confirm_password:   str,
        current_session_id: str = "",
    ) -> AuthResult:
        user = self._repo.get_by_id(user_id)
        if not user:
            return AuthResult.fail("User not found.", "NOT_FOUND")

        if not self._pw.verify(current_password, user.password_hash):
            return AuthResult.fail("Current password is incorrect.", "INVALID_CREDENTIALS")

        errors = self._pw.validate_strength(new_password)
        if errors:
            return AuthResult.fail(errors[0], "WEAK_PASSWORD")
        if new_password != confirm_password:
            return AuthResult.fail("Passwords do not match.", "PASSWORD_MISMATCH")

        new_hash = self._pw.hash(new_password)
        self._repo.update_password(user.id, new_hash)
        # Revoke all other sessions so stolen tokens cannot be used after a password change.
        if current_session_id:
            self._sessions.revoke_others(user.id, current_session_id)
        else:
            self._sessions.revoke_all(user.id)
        self._audit(user.email, "password_changed")
        return AuthResult.ok(user)

    # ── Profile ───────────────────────────────────────────────────────────────

    def update_profile(self, user_id: str, full_name: str) -> AuthResult:
        errors = validate_name(full_name)
        if errors:
            return AuthResult.fail(errors[0], "VALIDATION_ERROR")
        self._repo.update_name(user_id, full_name.strip())
        user = self._repo.get_by_id(user_id)
        return AuthResult.ok(user)

    def get_user(self, user_id: str) -> Optional[User]:
        return self._repo.get_by_id(user_id)

    def get_user_by_email(self, email: str) -> Optional[User]:
        return self._repo.get_by_email(normalize_email(email))

    # ── Sessions ──────────────────────────────────────────────────────────────

    def validate_session(self, session_id: str) -> Optional[User]:
        """Return User if session is valid, else None."""
        session = self._sessions.validate(session_id)
        if not session:
            return None
        user = self._repo.get_by_id(session.user_id)
        if not user or not user.is_active():
            return None
        return user

    def get_active_sessions(self, user_id: str) -> list:
        return self._sessions.list_active(user_id)

    def revoke_session(self, session_id: str, user_id: str) -> bool:
        session = self._sessions.get(session_id)
        if not session or session.user_id != user_id:
            return False
        self._sessions.revoke(session_id)
        return True

    # ── Admin operations (owner/admin role only — caller must enforce) ────────

    def admin_verify_user(self, target_user_id: str, admin_email: str = "") -> AuthResult:
        """Manually verify a user's email. Owner/admin use only."""
        guard = self._verify_admin_caller(admin_email)
        if guard is not None:
            return guard
        user = self._repo.get_by_id(target_user_id)
        if not user:
            return AuthResult.fail("User not found.", "NOT_FOUND")
        self._repo.verify_email(target_user_id)
        self._audit(admin_email, "admin_manual_verify", resource=user.email)
        user = self._repo.get_by_id(target_user_id)
        return AuthResult.ok(user)

    def admin_update_role(self, target_user_id: str, new_role: str, admin_email: str = "") -> AuthResult:
        guard = self._verify_admin_caller(admin_email)
        if guard is not None:
            return guard

        # Validate the requested role before hitting the DB
        if new_role not in _VALID_ROLES:
            return AuthResult.fail(
                f"'{new_role}' is not a valid assignable role.",
                "VALIDATION_ERROR",
            )

        user = self._repo.get_by_id(target_user_id)
        if not user:
            return AuthResult.fail("User not found.", "NOT_FOUND")

        old_role = user.role

        # Hard protection: owner role is permanent — cannot be changed.
        if old_role == "owner":
            return AuthResult.fail(
                "The owner role is protected and cannot be changed.",
                "FORBIDDEN",
            )

        self._repo.update_role(target_user_id, new_role)
        self._audit(
            admin_email, "admin_role_change",
            resource=user.email,
            detail=f"{old_role} → {new_role}",
        )
        return AuthResult.ok(self._repo.get_by_id(target_user_id))

    def admin_set_status(self, target_user_id: str, status: AccountStatus, admin_email: str = "") -> AuthResult:
        guard = self._verify_admin_caller(admin_email)
        if guard is not None:
            return guard
        user = self._repo.get_by_id(target_user_id)
        if not user:
            return AuthResult.fail("User not found.", "NOT_FOUND")
        self._repo.update_status(target_user_id, status)
        self._audit(admin_email, f"admin_set_status_{status.value}", resource=user.email)
        return AuthResult.ok(self._repo.get_by_id(target_user_id))

    def admin_list_users(self) -> list:
        return self._repo.list_all()

    # ── Internal helpers ──────────────────────────────────────────────────────

    def _verify_admin_caller(self, admin_email: str) -> "AuthResult | None":
        """Return None if the caller has owner/admin role; return AuthResult.fail() otherwise."""
        if not admin_email:
            return AuthResult.fail("Admin email is required for this operation.", "FORBIDDEN")
        caller = self._repo.get_by_email(normalize_email(admin_email))
        if caller and caller.role in ("owner", "admin") and caller.is_active():
            return None
        return AuthResult.fail(
            "Caller does not have sufficient privileges (owner or admin required).",
            "FORBIDDEN",
        )

    # ── Internal ──────────────────────────────────────────────────────────────

    def _send_verification(self, user: User) -> str | None:
        """Send verification email. Returns raw token (for in-app link), or None on failure."""
        try:
            raw_token  = self._tok.generate()
            token_hash = self._tok.hash(raw_token)
            self._repo.set_verification_token(
                user.id, token_hash, time.time() + _VERIFICATION_TTL
            )
            svc = _get_email_service()
            svc.send_verification_email(user.email, user.full_name, raw_token)
            self._audit(user.email, "verification_email_sent")
            return raw_token
        except Exception as exc:
            _log.error("verification email failed for %s: %s", user.email, exc)
            return None

    def _check_rate_limit(self, email: str) -> bool:
        try:
            from psydox.security.ratelimit import get_rate_limiter
            return get_rate_limiter().check(email, "login")
        except Exception as exc:
            _log.warning("rate limiter unavailable — denying login request: %s", exc)
            return False  # fail closed

    def _audit(self, email: str, action: str, resource: str = "", detail: str = "") -> None:
        try:
            from psydox.security.audit import get_audit_log
            get_audit_log().log(email, action, resource, detail)
        except Exception:
            pass

    def _db(self):
        from psydox.storage.database import get_db
        return get_db()


def _get_email_service():
    from psydox.auth.email import get_email_service
    return get_email_service()


def _is_console_email() -> bool:
    """True when no real email service is configured (console/dev mode)."""
    from psydox.auth.email import ConsoleEmailService
    return isinstance(_get_email_service(), ConsoleEmailService)


# ── Registration Excel log ────────────────────────────────────────────────────

import threading as _threading
_reg_log_lock = _threading.Lock()

def _append_registration_log(user: "User") -> None:
    """Append one row to user_registrations.xlsx next to the DB file.
    Thread-safe.  Non-fatal on any error.
    Password hashes are NEVER written to the log.
    """
    try:
        import datetime, openpyxl
        from psydox.storage.database import get_db_path

        log_path = get_db_path().parent / "user_registrations.xlsx"

        with _reg_log_lock:
            if log_path.exists():
                wb = openpyxl.load_workbook(log_path)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Registrations"
                ws.append([
                    "Full Name", "Email", "Role", "Status", "Registered At",
                ])
                from openpyxl.styles import Font, PatternFill
                hdr_fill = PatternFill("solid", fgColor="FF6600")
                hdr_font = Font(bold=True, color="FFFFFF")
                for cell in ws[1]:
                    cell.fill = hdr_fill
                    cell.font = hdr_font

            registered_at = datetime.datetime.fromtimestamp(
                user.created_at
            ).strftime("%Y-%m-%d %H:%M:%S")

            ws.append([
                user.full_name,
                user.email,
                user.role,
                user.status.value,
                registered_at,
            ])
            wb.save(log_path)
    except Exception as exc:
        _log.warning("registration log write failed (non-fatal): %s", exc)


def get_registration_log_excel() -> bytes | None:
    """Return the current user_registrations.xlsx as bytes, or None if empty."""
    try:
        from psydox.storage.database import get_db_path
        log_path = get_db_path().parent / "user_registrations.xlsx"
        if log_path.exists():
            return log_path.read_bytes()
    except Exception:
        pass
    return None


def get_all_users_excel() -> bytes:
    """Generate a fresh Users Excel from the DB (always up-to-date).
    Password hashes are NEVER included in exports.
    """
    import io, datetime, openpyxl
    from psydox.auth.repository import get_user_repository
    from openpyxl.styles import Font, PatternFill

    repo = get_user_repository()
    users = repo.list_all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"
    headers = ["Full Name", "Email", "Role", "Status", "Registered At",
               "Last Login", "Email Verified"]
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="FF6600")
    hdr_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font

    for u in users:
        def _fmt(ts):
            if not ts:
                return ""
            return datetime.datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
        ws.append([
            u.full_name,
            u.email,
            u.role,
            u.status.value,
            _fmt(u.created_at),
            _fmt(u.last_login_at),
            "Yes" if u.email_verified else "No",
        ])

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_service: AuthService | None = None
_service_lock = None


def get_auth_service() -> AuthService:
    """Process-level singleton — thread-safe via DB."""
    global _service, _service_lock
    if _service_lock is None:
        import threading
        _service_lock = threading.Lock()
    if _service is None:
        with _service_lock:
            if _service is None:
                _service = AuthService()
    return _service
