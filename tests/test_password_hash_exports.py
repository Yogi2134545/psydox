"""
Regression tests: password_hash must never appear in any exported output.
These tests verify that Excel/JSON exports produced by auth.service do not
leak the bcrypt hash stored in the users table.
"""
import io
import pytest


def _get_all_users_excel_bytes(db_path: str) -> bytes:
    """
    Point PSYDOX_DB_PATH at a fresh temp DB, let init_db() create the full
    schema, insert one user with a real bcrypt hash, then call get_all_users_excel().
    """
    import os
    import bcrypt

    os.environ["PSYDOX_DB_PATH"] = db_path
    # Patch the module-level cached path so get_db() uses our test DB.
    import psydox.storage.database as _dbmod
    _dbmod._DB_PATH = db_path

    # Trigger full schema creation via the real init_db()
    db = _dbmod.get_db()

    # Seed one user with a real bcrypt hash
    phash = bcrypt.hashpw(b"SuperSecret123!", bcrypt.gensalt()).decode()
    import time
    now = time.time()
    db.execute(
        """INSERT OR IGNORE INTO users
           (id, email, full_name, role, status, password_hash,
            email_verified, created_at, updated_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        ("u-test", "test@example.com", "Test User", "admin", "active",
         phash, 1, now, now),
    )
    db.commit()

    from psydox.auth.service import get_all_users_excel
    return get_all_users_excel()


def test_excel_export_no_password_hash(tmp_path):
    db_file = str(tmp_path / "test.db")
    excel_bytes = _get_all_users_excel_bytes(db_file)
    assert excel_bytes is not None and len(excel_bytes) > 0

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active

    all_text = []
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is not None:
                all_text.append(str(cell).lower())

    combined = " ".join(all_text)
    assert "password" not in combined, "Exported Excel must not contain any 'password' column"
    assert "$2b$" not in combined, "Exported Excel must not contain bcrypt hash"
    assert "$2a$" not in combined, "Exported Excel must not contain bcrypt hash"


def test_excel_export_expected_columns(tmp_path):
    db_file = str(tmp_path / "test2.db")
    excel_bytes = _get_all_users_excel_bytes(db_file)

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active

    headers = [str(c).lower() for c in next(ws.iter_rows(values_only=True))]
    assert any("email" in h for h in headers)
    assert any("role" in h for h in headers)
    assert not any("hash" in h for h in headers)
    assert not any("password" in h for h in headers)


class TestRetryReasonClassification:
    """Grouped here as a smoke-test for _is_retryable_reason since it governs security."""

    def test_timeout_is_retryable(self):
        from psydox.batch.processor import _is_retryable_reason
        assert _is_retryable_reason("Connection timeout after 15s")

    def test_http_429_is_retryable(self):
        from psydox.batch.processor import _is_retryable_reason
        assert _is_retryable_reason("HTTP 429 Too Many Requests")

    def test_http_500_is_retryable(self):
        from psydox.batch.processor import _is_retryable_reason
        assert _is_retryable_reason("HTTP 500 Internal Server Error")

    def test_http_404_is_permanent(self):
        from psydox.batch.processor import _is_retryable_reason
        assert not _is_retryable_reason("HTTP 404 Not Found")

    def test_http_403_is_permanent(self):
        from psydox.batch.processor import _is_retryable_reason
        assert not _is_retryable_reason("HTTP 403 Forbidden")

    def test_empty_reason_is_permanent(self):
        from psydox.batch.processor import _is_retryable_reason
        assert not _is_retryable_reason("")
